"""
Обработка скачанных файлов, запись данных в основную таблицу и/или отчеты
"""
import os
import re
import shutil

from datetime import datetime as dt

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from message_for_user import MessageError as Me


class BasicTable:
    """
    Запись данных в основную таблицу и при необходимости в отчеты.
    """

    def __init__(self, paths: dict):
        self.dir_path = 'email_files'  # путь к скачанным с почты файлам
        self.basic_table_path = paths["Database"]  # путь к основной таблице pearl.xlsx
        self.reports_path = paths["Reports"]  # путь к файлу отчетов
        self.contractors_path = paths["Contractors"]  # путь к файлу контрагентов
        self.column_names_list = list()  # список имен столбцов
        self.downloaded_files = list()  # список скачанных файлов
        self.data_for_write_dict = dict()  # словарь данных для записи в основную таблицу
        self.distributor_color = dict()  # словарь дистрибьютеров с указанием цвета для закрашивания ячейки
        self.color = None  # цвет для окрашивания ячейки
        self.report = str()  # текст для записи в отчет
        self.flag = False  # флаг необходимости записи в отчеты
        self.VAR_COLUMN_NAMES = {  # возможные варианты названия столбцов
            "Year": ["Год", ],
            "Month": ["Месяц", ],
            "Date": ["Дата", ]
        }

        self.wb_basic_table = load_workbook(self.basic_table_path)
        self.basic_table = self.wb_basic_table.active
        self.wb_report = load_workbook(self.reports_path)
        self.reports_table = self.wb_report.active
        self.wb_contractors = load_workbook(self.basic_table_path)
        self.contractor_table = self.wb_contractors.active

        self.run()

    def run(self) -> None:
        """
        Определение порядка запуска методов.
        :return: None
        """
        self.form_column_names_dict()
        self.work_with_downloaded_files()
        self.wb_basic_table.save(self.basic_table_path)
        self.wb_report.save(self.reports_path)
        self.wb_contractors.save(self.contractors_path)

    def form_column_names_dict(self) -> None:
        """
        Формирование словаря с названиями столбцов таблицы кроме столбцов ИНН и Контрагент
         (по ним отдельная расширенная проверка).
        Далее копия этого словаря будет использоваться при формировании набора данных
         для записи в основную таблицу.
        :return: None
        """
        self.column_names_list = [self.basic_table.cell(row=1, column=j).value
                                  for j in range(1, self.basic_table.max_column + 1)]
        self.data_for_write_dict = {key: None for key in self.column_names_list}
        self.data_for_write_dict.pop("ИНН")
        self.data_for_write_dict.pop("Контрагент")

    def work_with_downloaded_files(self) -> None:
        """
        Перебираются все скачанные файлы.
        Для каждого файла запускается метод формирования набора данных
         для записи в основную таблицу (form_data_for_basic_table).
        И формируется словарь, на основании которого
         будет производиться закрашивание файла Дистрибьютеры.xlsx
        :return: None
        """
        self.downloaded_files = list(os.walk(self.dir_path))[0][2]
        for file in self.downloaded_files:
            self.color = "FF92D050"  # светло-зелёный
            self.form_data_for_basic_table(file)
            self.distributor_color[file[:file.rfind(".")]] = self.color

    def form_data_for_basic_table(self, source_file: str) -> None:
        """
        Формируется словарь с данными, которые необходимо записать в основную таблицу и/или в отчеты,
         и вызывается метод, записывающий эти данные.
        Если файл с данными полностью корректен, то запись производится только в основную таблицу.
        В случае отсутствия столбца ИНН или некорректном значении ИНН,
         запись производится только в файл отчетов.
        В случае отсутствия каких-нибудь других необходимых данных,
         запись в основную таблицу производится,
          но при этом также производится запись в файл отчетов.
        :param source_file: Строка с именем файла.
        :return: None
        """
        wb_source_file = load_workbook(os.path.join(self.dir_path, source_file))
        table = wb_source_file.active
        column_names = [table.cell(row=1, column=j).value.upper() for j in range(1, table.max_column + 1)]

        if "ИНН" in column_names:

            for i in range(2, table.max_row + 1):
                data_for_write = self.data_for_write_dict.copy()
                column_names = self.check_var_column_names([name.upper() for name in column_names])
                for column in data_for_write:
                    try:
                        data_for_write[column] = table.cell(row=i, column=column_names.index(column.upper()) + 1).value
                    except ValueError:
                        data_for_write[column] = '-----'
                        self.report += f"Отсутствует столбец '{column}'"
                        self.flag = True
                        print(self.report)

                itn = str(table.cell(row=i, column=column_names.index("ИНН") + 1).value)
                if re.fullmatch(r'\d{10}', itn) or re.fullmatch(r'\d{12}', itn):
                    data_for_write["ИНН"] = itn
                    data_for_write["Контрагент"] = self.get_contractor_name(itn)
                    self.write_to_basic_table(data_for_write)
                    if self.flag:
                        self.write_to_reports(data_for_write)
                else:
                    self.color = "FFFFC000"  # оранжевый
                    data_for_write["ИНН"] = "-----"
                    data_for_write["Контрагент"] = "-----"
                    self.report = "Некорректное значение ИНН"
                    self.write_to_reports(data_for_write)

        else:
            self.color = "FFFF0000"  # красный
            self.report = "Отсутствует столбец ИНН"
            self.write_to_reports(self.data_for_write_dict)

    def check_var_column_names(self, column_names: list) -> list:
        """
        Проверка имен столбцов на допустимые вариации.
        :param column_names: Список имен столбцов
        :return: Обновленный список имен столбцов
        """
        for name in self.VAR_COLUMN_NAMES:
            if name not in column_names:
                for similarly in self.VAR_COLUMN_NAMES[name]:
                    if similarly in column_names:
                        column_names[column_names.index(similarly)] = name
        return column_names

    def write_to_basic_table(self, data: dict) -> None:
        """
        Производится запись данных в основную таблицу в самый конец.
        :param data: Словарь данных для записи в основную таблицу (ключи - названия столбцов).
        :return: None
        """
        start_row = self.basic_table.max_row
        for j in range(1, len(self.column_names_list) + 1):
            self.basic_table.cell(row=start_row + 1, column=j).value = data[self.column_names_list[j - 1]]

    def write_to_reports(self, data: dict) -> None:
        """
        Производится запись данных в отчеты в самый конец.
        :param data: Словарь данных для записи в отчет. Ключами выступают названия столбцов.
        :return:
        """
        row = self.reports_table.max_row + 1
        self.reports_table.cell(row=row, column=1).value = self.report
        for j in range(1, len(self.column_names_list) + 1):
            self.reports_table.cell(row=row, column=j + 1).value = data[self.column_names_list[j - 1]]
            self.reports_table.cell(row=row, column=1).fill = PatternFill(fgColor=self.color, fill_type="solid")
        self.report = ""
        self.flag = False

    def get_contractor_name(self, itn: str) -> str:
        """
        Получение правильного имени контрагента по его ИНН из файла данных.
        :param itn: ИНН контрагента в строчном формате.
        :return: Имя контрагента.
        """
        # Поиск индексов необходимых столбцов
        index_itn = False
        index_name = False
        for j in range(1, self.contractor_table.max_column + 1):
            if not index_itn and self.contractor_table.cell(row=1, column=j).value == "ИНН":
                index_itn = j
            elif not index_name and self.contractor_table.cell(row=1, column=j).value == "Контрагент":
                index_name = j
            if index_name and index_itn:
                break

        for i in range(2, self.contractor_table.max_row + 1):
            if str(self.contractor_table.cell(row=i, column=index_itn).value) == itn:
                return self.contractor_table.cell(row=i, column=index_name).value

        self.report = "Обнаружен новый контрагент; "
        self.flag = True
        return "-----"

    @staticmethod
    def form_report_for_sr(paths: dict, status_data: dict) -> None:
        """
        Проверка окончания отчетности в текущем месяце,
         перенос файлов для торговых представителей в архив
          и создание новых актуальных для файлов.
        :param paths: Словарь с путями к ключевым файлам
        :param status_data: Словарь статусов
        :return: None
        """
        if status_data["доклад без замечаний"] != status_data["Всего"]:
            return None

        src = paths["sales_representatives"]
        dst = paths["sales_representatives_archive"]
        cart = dict()
        for file in list(os.walk(src))[0][2]:  # перенос всех докладов в архив
            filename = f"{file}_{dt.now().year}_{dt.now().month}.xlsx"
            shutil.move(os.path.join(src, file), os.path.join(dst, filename))

        wb_basic_table = load_workbook(paths["Database"])
        basic_table = wb_basic_table.active

        col = 0
        for j in range(1, basic_table.max_column + 1):
            if basic_table.cell(row=1, column=j).value == "KAM":
                col = j
                break

        if col == 0:
            Me.message_window("В основной таблице не столбца 'КАМ'")

        for i in range(2, basic_table.max_row + 1):
            row = [cell.value for cell in basic_table[i]]
            value = basic_table.cell(row=i, column=col).value
            try:
                cart[value].append(row)
            except KeyError:
                cart[value] = [row, ]

        for s_r in cart:  # s_r - sales representatives
            wb = Workbook()
            ws = wb.active
            for row in cart[s_r]:
                ws.append(row)
            wb.save(os.path.join(src, f"{s_r}.xlsx"))
        return None  # Чтобы pylint не ругался


class OtherFiles:
    """
    Для работы с файлами, не требующими сложной обработки.
    """

    @staticmethod
    def sales_representatives(path: str, filename: str) -> str or None:
        """
        Получение email из базы торговых представителей по его имени (имя файла).
        :param path: Путь к базе торговых представителей
        :param filename: Имя файла
        :return: Строка с email или None
        """
        wb = load_workbook(path)
        ws = wb.active
        r_s = filename[:filename.rfind(".")]
        for row in ws:
            if row[0].value == r_s:
                return row[1].value

        note = [r_s, "необходимо добавить"]
        ws.append(note)
        row = ws.max_row
        for j in range(1, 3):
            ws.cell(row=row, column=j).fill = PatternFill(fgColor="FFFF0000", fill_type="solid")
        return None
