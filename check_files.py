"""
Проверка файлов
"""
import os.path
import json
# install openpyxl, pandas
import datetime as dt
from openpyxl import load_workbook

from message_error import MessageError as Me


class CheckFiles:
    """
    Получение путей и их проверка
    """

    def __init__(self):
        self.paths = None
        self.get_paths()
        self.check_files()

    def get_paths(self) -> None:
        """
        Получение путей к файлам с данными из служебного файла Data.txt.
        При неудаче выводится окно с ошибкой
        :return: None
        """
        try:
            with open('Data.txt', 'r', encoding='utf-8') as file:
                self.paths = json.load(file)
        except json.decoder.JSONDecodeError:
            Me.message_window('Некорретный формат файла Data.txt')
        except FileNotFoundError:
            Me.message_window('Файл Data.txt не найден')
        except Exception as except_:
            print(except_)
            Me.message_window('Что-то пошло не так!!!')

    def check_files(self) -> None:
        """
        Производится проверка наличия файлов по указанным путям.
        :return: None
        """
        for file in self.paths:
            if not os.path.exists(self.paths[file]):
                Me.message_window(f'Файл {file} не найден')


class Distributors:
    """
    Работа с файлом дистрибьютеров
    """

    def __init__(self, path: str):
        self.date_in_file = None
        self.month = None
        self.debtors = []
        # self.workbook = None
        self.path = path
        self.wb_distributors = None
        self.distributors_table = None

        self.get_cell_a1()
        self.is_valid_data_cell()
        self.check_month_in_file()
        self.get_debtors()

    def get_cell_a1(self) -> None:
        """
        Получение содержимого ячейки А1 из файла формата xlsx.
        :return: None
        """
        self.wb_distributors = load_workbook(self.path)
        self.distributors_table = self.wb_distributors.active
        self.date_in_file = self.distributors_table.cell(row=1, column=1)

    def is_valid_data_cell(self) -> None:
        """
        Проверка первой ячейки таблицы. Содержимое ячейки должно быть в формате даты.
        :return: None
        """
        if not isinstance(self.date_in_file.value, dt.datetime):
            Me.message_window('В файле Дистрибьютеры.xlsx в ячейке А1 '
                              'отсутствует дата в нужном формате (ДД.ММ.ГГГГ).')

    def get_debtors(self) -> None:
        """
        Формирование списка "должников".
        Критерием является незакрашенность ячейки с именем дистрибьютера.
        :return: None
        """
        for i in range(2, self.distributors_table.max_row + 1):
            if self.distributors_table.cell(row=i, column=1).fill.fgColor.value in ["00FFFFFF", "00000000", 0, "FFFFFFFF"]:
                self.debtors.append(self.distributors_table.cell(row=i, column=2).value)

    def check_month_in_file(self) -> None:
        """
        Сравнение указанного в файле месяца с текущим.
        Если не совпадает, то закрасить весь файл в белый
        :return: None
        """
        current_month = dt.date.today().month
        month_in_file = self.date_in_file.value.month
        if month_in_file != current_month:
            # wb_distributors = load_workbook(self.path)
            # distributors_table = wb_distributors.active

            for cell_row in self.distributors_table["A2": f"A{self.distributors_table.max_row + 1}"]:
                # print(cell_row[0])
                cell_row[0].fill.fgColor.value = '00FFFFFF'

            self.wb_distributors.save(self.path)

    @staticmethod
    def set_month_in_file(path) -> None:
        """
        Запись текущей даты в файл.
        :return: None
        """
        wb_distributor = load_workbook(path)
        distributor_table = wb_distributor.active
        distributor_table.cell(row=1, column=1).value = dt.datetime.today()
        wb_distributor.save(path)
