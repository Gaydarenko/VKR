"""
Проверка файлов
"""
import os.path
import json
# install openpyxl, pandas
import datetime as dt
import openpyxl

from message_error import MessageError as Me


class CheckFiles:

    def __init__(self):
        self.paths = None
        self.get_paths()
        self.check_files()

    def get_paths(self) -> None:
        """
        Получение путей к файлам с данными из файла служебного файла Data.txt.
        При неудаче выводится окно с ошибкой
        :return: None
        """
        try:
            with open('Data.txt', 'r', encoding='utf-8') as file:
                self.paths = json.load(file)
        except json.decoder.JSONDecodeError:
            Me.message_window('Некорретный формат файла Data.txt')
        except FileNotFoundError:
            Me.message_window('Файл Data.txt не найден')
        except Exception as except_:
            print(except_)
            Me.message_window('Что-то пошло не так!!!')

    def check_files(self) -> None:
        """
        Производится проверка наличия файлов по указанным путям.
        :return: None
        """
        for file in self.paths:
            if not os.path.exists(self.paths[file]):
                Me.message_window(f'Файл {file} не найден')


class Distributors:

    def __init__(self, path: str):
        self.cell = None
        self.month = None
        self.debtors = []
        self.workbook = None
        self.path = path
        self.get_cell_a1()
        self.is_valid_data_cell()
        self.check_month_in_file()
        self.get_debtors()

    def get_cell_a1(self) -> None:
        """
        Получение содержимого ячейки А1 из файла формата xlsx.
        :return: None
        """
        self.workbook = openpyxl.load_workbook(self.path)
        table = self.workbook.active
        self.cell = table.cell(row=1, column=1)

    def is_valid_data_cell(self) -> None:
        """
        Проверка первой ячейки таблицы. Содержимое ячейки должно быть в формате даты.
        :return: None
        """
        if not isinstance(self.cell.value, dt.datetime):
            Me.message_window('В файле Дистрибьютеры.xlsx в ячейке А1 '
                              f'отсутствует дата в нужном формате (ДД.ММ.ГГГГ).')

    def get_debtors(self) -> None:
        """
        Формирование списка "должников".
        Критерием является незакрашенность ячейки с именем дистрибьютера.
        :return: None
        """
        for row in self.workbook["Sheet"].iter_rows(min_row=2, min_col=2, max_col=2):
            # print(f"{row[0].fill.fgColor.value} - {row[0].value}")
            if row[0].fill.fgColor.value in ["00FFFFFF", "00000000", 0]:
                self.debtors.append(row[0].value)

    def check_month_in_file(self) -> None:
        """
        Сравнение указанного в файле месяца с текущим.
        Если не совпадает, то закрасить весь файл в белый
        :return: None
        """
        current_month = dt.date.today().month
        month_in_file = self.cell.value.month
        if month_in_file != current_month:
            wb = openpyxl.load_workbook(self.path)
            sheet = wb[wb.sheetnames[0]]    # получение первого листа по его имени
            rows = sheet.max_row
            sheet_obj = wb.active
            for cell_row in sheet_obj["A2": f"A{rows+1}"]:
                cell_row[0].fill.fgColor.value = '00FFFFFF'
            wb.save(self.path)

    def set_month_in_file(self) -> None:
        """
        Запись текущей даты в файл.
        :return: None
        """
        self.cell.value = dt.datetime.today()
        self.workbook.save(self.path)
