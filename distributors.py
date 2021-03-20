"""
Работа с файлом дистрибьютеров
"""
import os
import shutil
import datetime as dt

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from message_for_user import MessageError as Me


class Distributors:
    """
    Работа с файлом дистрибьютеров
    """

    def __init__(self, path: dict):
        self.date_in_file = None
        self.month = None
        self.debtors = []
        self.path = path["Distributors"]
        self.basic_table_path = path["Database"]
        self.basic_table_archive_path = path["basic_tables_archive"]
        self.wb_distributors = load_workbook(self.path)
        self.distributors_table = self.wb_distributors.active

        self.basic_run()

    def basic_run(self) -> None:
        """
        Порядок запуска методов при основном обращении
        :return: None
        """
        self.get_cell_a1()
        self.is_valid_data_cell()
        self.check_month_in_file()
        self.get_debtors()

    def get_cell_a1(self) -> None:
        """
        Получение содержимого ячейки А1 из файла формата xlsx.
        :return: None
        """
        self.date_in_file = self.distributors_table.cell(row=1, column=1)

    def is_valid_data_cell(self) -> None:
        """
        Проверка первой ячейки таблицы. Содержимое ячейки должно быть в формате даты.
        :return: None
        """
        if not isinstance(self.date_in_file.value, dt.datetime):
            Me.message_window('В файле Дистрибьютеры.xlsx в ячейке А1 '
                              'отсутствует дата в нужном формате (ДД.ММ.ГГГГ).')

    def check_month_in_file(self) -> None:
        """
        Сравнение указанного в файле месяца с текущим.
        Если не совпадает, то основная таблица копируется в архив
         и весь файл дистрибьюторов закрашивается в белый.
        :return: None
        """
        today = dt.date.today()
        dt_current_month = dt.datetime(today.year, today.month, 1)
        if self.date_in_file.value < dt_current_month:
            self.basic_table_to_archive()

            for i in range(2, self.distributors_table.max_row + 1):
                for j in range(1, self.distributors_table.max_column + 1):
                    self.distributors_table.cell(row=i, column=j).fill.fgColor.value = "00FFFFFF"

            self.wb_distributors.save(self.path)

    def get_debtors(self) -> None:
        """
        Формирование списка "должников".
        Критерием является незакрашенность ячейки с именем дистрибьютера.
        :return: None
        """
        for i in range(2, self.distributors_table.max_row + 1):
            if self.distributors_table.cell(row=i, column=1).fill.fgColor.value in ["00FFFFFF", "00000000", 0,
                                                                                    "FFFFFFFF", "FFFF0000"]:
                self.debtors.append(self.distributors_table.cell(row=i, column=2).value)

    @staticmethod
    def set_month_in_file(path) -> None:
        """
        Запись текущей даты в файл.
        :param path: путь к файлу с таблицей дистрибьютеров
        :return: None
        """
        wb_distributor = load_workbook(path)
        distributor_table = wb_distributor.active
        distributor_table.cell(row=1, column=1).value = dt.datetime.today()
        wb_distributor.save(path)

    @staticmethod
    def form_status_data(path: str) -> dict:
        """
        Формирование словаря с данными о ходе выполнения общей задачи.
        :param path: путь к файлу с таблицей дистрибьютеров
        :return: словарь с данными
        """
        statuses = ["не прислали доклад",
                    "доклад без замечаний",
                    "требует участия человека",
                    "некорректных докладов",
                    "невозможно определить статус",
                    "Всего"]
        progress = {status: 0 for status in statuses}

        wb_distributor = load_workbook(path)
        distributor_table = wb_distributor.active

        for i in range(2, distributor_table.max_row + 1):
            color = distributor_table.cell(row=i, column=1).fill.fgColor.value
            if color in ["00FFFFFF", "00000000", 0, "FFFFFFFF"]:  # белый цвет
                id_status = 0
            elif color in ["FF92D050", ]:  # светло-зелёный
                id_status = 1
            elif color in ["FFFFC000", ]:  # оранжевый
                id_status = 2
            elif color in ["FFFF0000", ]:  # красный
                id_status = 3
            else:
                id_status = 4
            progress[statuses[id_status]] += 1

        progress[statuses[5]] = distributor_table.max_row - 1
        return progress

    def basic_table_to_archive(self) -> None:
        """
        Копирование файла основной таблицы в папку с архивными версиями
        :return: None
        """
        filename = f"pearl_{self.date_in_file.value.year}_{self.date_in_file.value.month}.xlsx"
        path = os.path.join(self.basic_table_archive_path, filename)
        shutil.copyfile(self.basic_table_path, path)

    @staticmethod
    def coloring(path: str, colors: dict) -> None:
        """
        Метод производит изменение цвета заливки для дистрибьютеров из словаря.
        :param path: путь к файлу с таблицей дистрибьютеров
        :param colors: Словарь цветового статуса для дистрибьютеров
        :return:
        """
        wb_distributor = load_workbook(path)
        distributor_table = wb_distributor.active
        for i in range(2, distributor_table.max_row + 1):
            distr = distributor_table.cell(row=i, column=2).value
            if distr in colors:
                for j in range(1, 3):
                    distributor_table.cell(row=i, column=j).fill = PatternFill(fgColor=colors[distr], fill_type="solid")
        wb_distributor.save(path)
